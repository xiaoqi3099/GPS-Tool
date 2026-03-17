"""
Excel文件处理模块
用于Android读取地址文件和生成报告
"""
import os


class ExcelHandler:
    """Excel文件处理器 - Android版本"""
    
    @staticmethod
    def read_addresses_from_file(file_path):
        """
        从文件读取地址列表
        支持格式: .xlsx, .xls, .txt, .csv
        """
        addresses = []
        
        try:
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext in ['.xlsx', '.xls']:
                # Excel文件处理
                try:
                    import openpyxl
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook.active
                    
                    # 读取B列地址
                    for cell in sheet['B'][1:]:
                        if cell.value:
                            addresses.append(str(cell.value))
                    
                    workbook.close()
                except ImportError:
                    # Android上可能没有openpyxl，尝试使用其他方式
                    addresses = ExcelHandler._read_excel_fallback(file_path)
            
            elif ext == '.txt':
                # 文本文件 - 每行一个地址
                with open(file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line:
                            addresses.append(line)
            
            elif ext == '.csv':
                # CSV文件 - 读取第二列
                import csv
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) >= 2 and row[1].strip():
                            addresses.append(row[1].strip())
                        elif row and row[0].strip():
                            addresses.append(row[0].strip())
            
        except Exception as e:
            print(f"读取文件错误: {e}")
        
        return addresses
    
    @staticmethod
    def _read_excel_fallback(file_path):
        """Excel回退方案 - 使用xlrd或pandas"""
        addresses = []
        
        try:
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            
            # 读取B列（索引1）
            for row in range(1, sheet.nrows):
                cell_value = sheet.cell_value(row, 1)
                if cell_value:
                    addresses.append(str(cell_value))
        except ImportError:
            try:
                import pandas as pd
                df = pd.read_excel(file_path)
                if 'B' in df.columns or len(df.columns) > 1:
                    col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
                    addresses = df[col].dropna().astype(str).tolist()
            except:
                pass
        
        return addresses
    
    @staticmethod
    def save_to_excel(data, output_path):
        """
        保存数据到Excel
        data: [{'address': str, 'lon': float, 'lat': float}, ...]
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "经纬度"
            
            # 设置列宽
            ws.column_dimensions['A'].width = 31
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            
            # 写入表头
            ws.append(["标准地址", "经度", "纬度"])
            
            # 设置表头居中
            for cell in ws["1:1"]:
                cell.alignment = Alignment(horizontal="center")
            
            # 用于存储已使用的坐标
            used_coordinates = set()
            
            # 写入数据
            for item in data:
                lon = "{:.6f}".format(item['lon'])
                lat = "{:.6f}".format(item['lat'])
                
                # 避免重复坐标
                while (lon, lat) in used_coordinates or lon.endswith('0') or lat.endswith('0'):
                    lon = "{:.6f}".format(float(lon) + 0.000001)
                    lat = "{:.6f}".format(float(lat) + 0.000001)
                
                used_coordinates.add((lon, lat))
                
                ws.append([item['address'], lon, lat])
                
                # 设置居中
                row = ws.max_row
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            
            wb.save(output_path)
            return True, output_path
            
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    def save_to_csv(data, output_path):
        """
        保存数据到CSV（更兼容Android）
        """
        try:
            import csv
            with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(["标准地址", "经度", "纬度"])
                
                for item in data:
                    writer.writerow([
                        item['address'],
                        "{:.6f}".format(item['lon']),
                        "{:.6f}".format(item['lat'])
                    ])
            
            return True, output_path
        except Exception as e:
            return False, str(e)
